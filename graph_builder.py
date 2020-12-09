import datetime  # noqa:D100
import logging
from collections import Counter
from pathlib import Path
from typing import Dict, List

import gmplot
import matplotlib.pyplot as plt
from openpyxl import Workbook
from xlsx2html import xlsx2html

from data_base_operation import get_hotels_coordinates, get_average_prices_for_city
from helper import set_lang_and_table_style

DATA_PATH = Path('Charts')
if not DATA_PATH.exists():
    logging.info('Making path for charts.')
    DATA_PATH.mkdir(exist_ok=True)


def schedule_quantity_rating(rating: List):
    """Build a histogram, where the hotel rating is horizontal, the count is vertical."""
    logging.info('Drawing histogram for scores.')
    plt.hist(rating, bins=100, rwidth=0.9, alpha=0.5, label='no', color='r')
    plt.title('Histogram of the number of hotels from their rating')
    plt.ylabel('Count of hotels')
    plt.xlabel('Hotel rating')
    fname = DATA_PATH / 'Number_of_hotels_by_rating'
    plt.savefig(fname)
    plt.close()
    logging.info('Histogram for scores drawn.')


def diagram_open_hotels(years):
    """Build a histogram of hotel registration on booking.com."""
    logging.info('Drawing years diagram.')
    years = sorted(years)
    count_year = Counter(years)
    years = []
    counts = []
    for year, count in count_year.items():
        years.append(int(year))
        counts.append(int(count))

    plt.bar(years, counts)
    plt.title('Hotel opening history histogram')
    plt.ylabel('Count of hotels')
    plt.xlabel('Opening year')
    fname = DATA_PATH / 'Number_of_hotels_by_year_of_registration_on_booking'
    plt.savefig(fname)
    plt.close()
    logging.info('Years diagram drawn.')


def pie_chart_from_scores(grouped_scores: Dict, city: str) -> None:
    """Draw pie chat of hotels scores."""
    logging.info(f'Drawing pie chart for {city}.')
    labels = 'Under 7', '7 to 8', '8 to 9', '9 to 10'
    amounts_of_scores = [len(grouped_scores['firstGroup']),
                         len(grouped_scores['secondGroup']),
                         len(grouped_scores['thirdGroup']),
                         len(grouped_scores['fourthGroup'])]
    total = sum(amounts_of_scores)
    fig, ax = plt.subplots()

    colors = ['#ffa88a', '#7b85cb', '#b491ca', '#b3d094' ]
    _, _, autotext = ax.pie(amounts_of_scores, colors=colors, autopct=lambda p: '{:,.0f}%'.format(p),
           wedgeprops={"edgecolor": "0", "linewidth": "0"})

    ax.axis('equal')

    plt.legend(
        title='Score groups',
        loc='upper left',
        labels=['%s, %s' % (
            label, score) for label, score in zip(labels, amounts_of_scores)],
        prop={'size': 10},
        bbox_to_anchor=(0.0, 1),
        bbox_transform=fig.transFigure
    )

    plt.title(f'{city}', loc='center')
    fname = DATA_PATH / f'Pie_chart_with_scores_for_{city}'
    plt.savefig(fname)
    plt.close()
    logging.info('Pie chart drawn.')


def get_table_of_ratio_data(ratio_data: Dict) -> None:
    """Get table with names of cities, hotels in each city, populations in each city and ratio."""
    wb = Workbook()
    ws = wb.active
    logging.info('Making table.')
    ws['A1'] = 'Город'
    ws['B1'] = 'Кол-во отелей'
    ws['C1'] = 'Население'
    ws['D1'] = 'Соотношение'

    for city_info in zip(ratio_data['cities'], ratio_data['hotels_amounts'],
                         ratio_data['cities_populations'], ratio_data['ratio']):
        city_info = list(city_info)
        ws.append(city_info)

    fname1 = DATA_PATH / 'ratio_data.xlsx'
    wb.save(fname1)

    html_table = xlsx2html(fname1)
    html_table.seek(0)
    html_table = html_table.read()
    fname2 = DATA_PATH / 'ratio_data.html'
    fname2.write_text(html_table)

    set_lang_and_table_style(fname2, "cp1251", "ru", "1", "5", "5",
                             "border: 1px solid black; font-size: 20.0px; height: 19px")
    logging.info('Table is done.')


def get_table_of_prices(cities: List) -> None:
    apartaments_prices = {}
    for city in cities:
        """Get tuple with minimal price, average minimal price, average price, average maximal price, maximal price for city"""
        apartaments_prices[city] = get_average_prices_for_city(city)
    wb = Workbook()
    ws = wb.active
    logging.info('Making table.')
    ws['A1'] = 'Город'
    ws['B1'] = 'Минимальная цена'
    ws['C1'] = 'Минимальная средняя цена'
    ws['D1'] = 'Средняя цена'
    ws['E1'] = 'Максимальная средняя цена'
    ws['F1'] = 'Максимальная цена'

    for city, value in apartaments_prices.items():
        prices = list(value)
        prices.insert(0, city)
        ws.append(prices)

    fname1 = DATA_PATH / 'prices_data.xlsx'
    wb.save(fname1)

    html_table = xlsx2html(fname1)
    html_table.seek(0)
    html_table = html_table.read()
    fname2 = DATA_PATH / 'prices_data.html'
    fname2.write_text(html_table)

    set_lang_and_table_style(fname2, "cp1251", "ru", "1", "5", "5",
                             "border: 1px solid black; font-size: 20.0px; height: 19px")
    logging.info('Table is done.')


def get_table_of_prices_by_star(cities: List) -> None:
    apartaments_prices = {}
    for city in cities:
        """Get tuple with minimal price, average minimal price, average price, average maximal price, maximal price for city"""
        apartaments_prices = get_average_prices_for_city(city, by_stars=True)
        wb = Workbook()
        ws = wb.active
        logging.info('Making table.')
        ws['A1'] = 'Звездность'
        ws['B1'] = 'Мин. цена'
        ws['C1'] = 'Макс. цена'
        ws['D1'] = 'Средняя цена'
        ws['E1'] = 'Нормированный диапазон от средней цены (%)'
        for star, information in apartaments_prices.items():
            if information:
                prices = [information['min_price'], information['max_price'], information['avg_price'], information['range']]
                prices.insert(0, star)
            else:
                prices = [star, 0, 0, 0, 0]
            ws.append(prices)

        fname1 = DATA_PATH / f'prices_data_by_star_{city}.xlsx'
        wb.save(fname1)

        html_table = xlsx2html(fname1)
        html_table.seek(0)
        html_table = html_table.read()
        fname2 = DATA_PATH / f'prices_data_by_star_{city}.html'
        fname2.write_text(html_table)

        set_lang_and_table_style(fname2, "cp1251", "ru", "1", "5", "5",
                                "border: 1px solid black; font-size: 20.0px; height: 19px")
        logging.info(f'Prices table by stars is done for {city}.')


def draw_map_by_coords(map_name: str) -> None:
    """Draw a map with labels at the given coordinates."""
    coordinates = get_hotels_coordinates()
    gmap = gmplot.GoogleMapPlotter(coordinates[0][1], coordinates[0][2], 5)
    logging.info(f"{datetime.datetime.now().strftime('%H:%M:%S')}:: Hotels: {len(coordinates)}")
    logging.info('Drawing map.')
    for hotel_coordinate in coordinates:
        hotel_name, latitude, longitude = hotel_coordinate
        try:
            latitude = float(latitude)
            longitude = float(longitude)
        except Exception:
            continue
        gmap.marker(latitude, longitude)
    fname = DATA_PATH / 'map_{0}.html'.format(map_name)
    gmap.draw(fname)
    logging.info('Map drawn.')
